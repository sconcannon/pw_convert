#! python3
# process_xl.py will open all excel files in a folder and place converted results in the same folder

import openpyxl,os
import transforms as tr
import settings
from datetime import datetime

# add the absolute path where your files to convert are located
# use double \\ for each slash
sourcePath = r'C:\Users\sconcannon\Documents\pw_file_convert\191222test'
os.chdir(sourcePath)
location = os.getcwd()

# get a timestamp for the result filenames
date = datetime.today()
timestamp = date.strftime("%Y%m%d%H%M")

# get a list of Excel files in the source location
fileList = [item for item in os.listdir(location) if item.endswith('.xlsx')]
mapping = settings.set_mapping()
season = settings.seasonName()
if fileList:
    os.mkdir(timestamp)

for file in fileList:
    # open each excel file in the sourcePath and create a new Excel file
    try:
        wb = openpyxl.load_workbook(file)
    except:
        print("WARNING Can't open ", file, " as an Excel workbook.")
    else:
        sheet = wb.get_active_sheet()
        resultFile = timestamp + "\\" + file + "_conv_" + timestamp + ".xlsx"
        sailsWb = openpyxl.Workbook()
        sailsSheet = sailsWb.active

        # set the SAILS sheet column headings according to the set_mapping dictionary
        for item in mapping:
            cell = item + '1'
            sailsSheet[cell] = mapping[item]
        for rownum in range(2, sheet.max_row +1):
            row=str(rownum)

            # set the B pubid and get the discount
            ew_pubId = sheet['P'+row].value
            pubDetails= tr.pub(ew_pubId)
            sailsSheet['B'+row].value = pubDetails['sails_code']

            # set the C format code
            ew_format = sheet['H'+row].value
            sailsSheet['C'+row].value = tr.book_format_encode(ew_format)

            # set the D release as frontlist (1)
            sailsSheet['D' + row].value = 1

            # set the E ISBN
            sailsSheet['E' + row].value = sheet['B' + row].value

            # set the F title
            ew_title = sheet['C' + row].value
            sailsTitle = tr.titleReformat(ew_title)
            sailsSheet['F' + row].value = sailsTitle

            # set the G subtitle
            sailsSheet['G' + row].value = sheet['D' + row].value

            # set the H series
            sailsSheet['H' + row].value = sheet['E' + row].value

            # set the I author
            sailsSheet['I' + row].value = sheet['F' + row].value

            # set the J author city
            # not available in the source
            sailsSheet['J' + row].value = ''

            #  set the K author state
            # not available in the source
            sailsSheet['K' + row].value = ''

            # set the L pub date mm/dd/YYYY
            sailsSheet['L' + row].value = tr.date_format(sheet['G' + row].value)

            # set the M discount ex 0.40
            sailsSheet['M' + row].value = pubDetails['disc']
            sailsSheet['M' + row].number_format = '0.00'

            # set the N usprice
            sailsSheet['N' + row].value = sheet['O' + row].value

            # set the O BISAC
            sailsSheet['O' + row].value = sheet['J' + row].value

            # set the P UPC
            # value not in source
            sailsSheet['P' + row].value = ''

            # set the Q trimsize
            sailsSheet['Q' + row].value = sheet['R' + row].value

            # set the R pagecnt
            sailsSheet['R' + row].value = sheet['S' + row].value

            # set the S printrun
            sailsSheet['S' + row].value = tr.printRun(sheet['T' + row].value)

            # set the T imprint
            sailsSheet['T' + row].value = sheet['Q' + row].value

            # set the U category
            sailsSheet['U' + row].value = sheet['K' + row].value

            # set the V catalogpage
            sailsSheet['V' + row].value = sheet['A' + row].value

            # set the W catalogcopy
            # not provided in the source
            sailsSheet['W' + row].value = ''

            # set the X local
            # not provided in the source
            sailsSheet['X' + row].value = ''

            # set the Y notes
            # not provided in the source
            sailsSheet['Y' + row].value = ''

            # set the Z catalogname
            sailsSheet['Z' + row].value = pubDetails['name'] +": " + season

            # set the AA ordering
            # not used
            sailsSheet['AA' + row].value = ''


        # save the resulting SAILS-formatted file
        sailsWb.save(resultFile)



